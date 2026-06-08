import openpyxl

file_path = r"d:\motor data\Motor List.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    print("Sheets in workbook:", wb.sheetnames)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")
        # Get first 5 rows to see what is filled
        rows = list(sheet.iter_rows(max_row=5, values_only=True))
        for r_idx, r in enumerate(rows, 1):
            # Print non-empty rows
            r_vals = [str(x) if x is not None else "" for x in r]
            if any(r_vals):
                print(f"  Row {r_idx}: {r_vals[:8]}")
except Exception as e:
    print("Error:", e)
