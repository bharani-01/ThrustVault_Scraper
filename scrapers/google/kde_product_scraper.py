"""
scrapers/google/kde_product_scraper.py
========================================
Scrapes a KDE Direct product page to get:
  1. Specs table (KV, resistance, weight, dimensions...)
  2. Performance data (from the linked Excel file on the page)
  3. Price, product name, links

KDE stores performance data as an Excel file (XLSX) linked from the page.
This scraper finds that link and parses all performance rows.
"""

import re
import io
import requests
from typing import Optional
from bs4 import BeautifulSoup
from utils.logger import get_logger

log = get_logger(__name__)

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.9',
}


def scrape_kde_product(url: str) -> tuple[list[dict], list[dict]]:
    """
    Scrape a single KDE Direct product page.

    Returns:
        motor_records  — list with one motor spec dict
        perf_points    — list of performance data row dicts
    """
    log.info(f'[kde_product] Scraping: {url}')

    try:
        r = requests.get(url, headers=HEADERS, timeout=20)
        r.raise_for_status()
    except Exception as e:
        log.warning(f'[kde_product] Fetch failed: {e}')
        return [], []

    soup = BeautifulSoup(r.text, 'html.parser')

    motor_name = _extract_motor_name(soup)
    specs      = _extract_specs_table(soup)
    price      = _extract_price(soup)
    xlsx_url   = _find_xlsx_url(soup, r.text)
    images     = _extract_images(soup)
    perf_pts   = _parse_xlsx(xlsx_url, motor_name) if xlsx_url else []

    motor_rec = {
        'motor_name':            motor_name,
        'company':               'KDE Direct',
        'kv_rating':             specs.get('kv_rating', ''),
        'internal_resistance':   specs.get('resistance', ''),
        'weight_g':              specs.get('weight', ''),
        'max_current':           specs.get('max_current', ''),
        'max_power':             specs.get('max_power', ''),
        'stator_size':           specs.get('stator_size', ''),
        'shaft_diameter':        specs.get('shaft', ''),
        'price':                 price,
        'link_motor':            url,
        'link_esc':              '',
        'link_propeller':        '',
        'xlsx_url':              xlsx_url or '',
        'images':                images,
        'perf_image':            next((i for i in images if 'kV' in i or 'perf' in i.lower() or 'kv' in i.lower()), ''),
        'source':                'kdedirect',
        'source_url':            url,
        **specs,
    }

    log.info(f'[kde_product] {motor_name}: {len(perf_pts)} perf points')
    return [motor_rec], perf_pts


# ── Internal helpers ──────────────────────────────────────────────────────────

def _extract_motor_name(soup: BeautifulSoup) -> str:
    for sel in ['h1.product-title', 'h1.product_name', 'h2.product-name', 'h1', '.product-title h1']:
        el = soup.select_one(sel)
        if el:
            return el.get_text(strip=True)

    # Fallback to og:title
    og = soup.find('meta', property='og:title')
    if og and og.get('content'):
        title = og.get('content').split('|')[0].strip()
        if 'Brushless Motor' in title:
            title = title.split('Brushless Motor')[0].strip()
        return title.strip()

    # Fallback to title tag
    if soup.title and soup.title.string:
        title = soup.title.string.split('|')[0].strip()
        return title.strip()

    return ''


def _extract_price(soup: BeautifulSoup) -> str:
    for sel in ['.product-single__price', '.price--regular', '.product__price', '[class*="price"]']:
        el = soup.select_one(sel)
        if el:
            txt = el.get_text(strip=True)
            if '$' in txt:
                return txt.split('\n')[0].strip()
    return ''


def _extract_specs_table(soup: BeautifulSoup) -> dict:
    """Parse the Specifications table on the KDE product page."""
    specs = {}

    # Find the #specifications div or any table with KV data
    for table in soup.find_all('table'):
        rows = table.find_all('tr')
        for row in rows:
            cells = [td.get_text(separator=' ', strip=True) for td in row.find_all(['td', 'th'])]
            if len(cells) < 2:
                continue
            key_raw = cells[0].lower()
            val     = cells[1]

            if 'kv' in key_raw and 'motor' in key_raw:
                specs['kv_rating'] = val.split()[0] if val else ''
            elif 'kv' in key_raw:
                specs['kv_rating'] = val.split()[0] if val else ''
            elif 'resistance' in key_raw:
                specs['resistance'] = val
            elif 'weight' in key_raw:
                specs['weight'] = re.sub(r'[^\d.]', '', val)
            elif 'max' in key_raw and 'current' in key_raw:
                specs['max_current'] = val
            elif 'max' in key_raw and 'power' in key_raw:
                specs['max_power'] = val
            elif 'stator' in key_raw and 'diam' in key_raw:
                specs['stator_size'] = val
            elif 'shaft' in key_raw:
                specs['shaft'] = val
            elif 'poles' in key_raw or 'magnets' in key_raw:
                specs['poles'] = val
            elif 'dimension' in key_raw or 'length' in key_raw:
                specs['dimensions'] = val

    return specs


def _extract_images(soup: BeautifulSoup) -> list[str]:
    """
    Collect all meaningful images from the page:
    - Performance chart (from #perfdata div)
    - Product gallery images
    Normalizes all URLs to https://
    """
    seen = set()
    images = []

    def add(src: str):
        if not src:
            return
        # Normalize protocol-relative URLs
        if src.startswith('//'):
            src = 'https:' + src
        if not src.startswith('http'):
            src = 'https://www.kdedirect.com' + src
        # Skip tiny icons / tracking pixels
        if any(x in src for x in ['icon', 'logo', 'badge', 'pixel', 'tracking', '1x1']):
            return
        if src not in seen:
            seen.add(src)
            images.append(src)

    # 1. Performance chart image (from #perfdata)
    perfdata = soup.find(id='perfdata')
    if perfdata:
        for img in perfdata.find_all('img', src=True):
            add(img['src'])
        for a in perfdata.find_all('a', href=True):
            href = a['href']
            if any(href.lower().endswith(ext) for ext in ['.png', '.jpg', '.jpeg', '.webp']):
                add(href)

    # 2. Product gallery images
    for img in soup.find_all('img', src=True):
        src = img.get('src', '') or img.get('data-src', '')
        # Only include CDN / product images (skip nav/cart icons)
        if any(d in src for d in ['cdn/shop', 'shopify.com/s/files', 'kdedirect.com/cdn']):
            # Skip tiny thumbnails (100px or less)
            if '_100x' in src or '_50x' in src or '_40x' in src:
                continue
            add(src)

    log.info(f'[kde_product] Found {len(images)} images')
    return images


def _find_xlsx_url(soup: BeautifulSoup, raw_html: str) -> Optional[str]:
    """Find the Excel performance data file linked on the page."""
    # Try <a> tags first
    for a in soup.find_all('a', href=True):
        href = a['href']
        if '.xlsx' in href.lower() or '.xls' in href.lower():
            if not href.startswith('http'):
                href = 'https:' + href if href.startswith('//') else 'https://www.kdedirect.com' + href
            log.info(f'[kde_product] Found XLSX: {href}')
            return href

    # Regex fallback in raw HTML
    m = re.search(r'https?://[^\s\'"<>]+\.xlsx[^\s\'"<>]*', raw_html)
    if m:
        return m.group()

    # CDN pattern
    m = re.search(r'//cdn\.shopify\.com/[^\s\'"<>]+\.xlsx[^\s\'"<>]*', raw_html)
    if m:
        url = 'https:' + m.group()
        log.info(f'[kde_product] Found XLSX (CDN): {url}')
        return url

    return None


def _parse_xlsx(xlsx_url: str, motor_name: str) -> list[dict]:
    """Download and parse the KDE Excel performance data file."""
    try:
        import openpyxl
    except ImportError:
        log.warning('[kde_product] openpyxl not installed — cannot parse XLSX')
        return []

    log.info(f'[kde_product] Downloading XLSX: {xlsx_url}')
    try:
        r = requests.get(xlsx_url, headers=HEADERS, timeout=30)
        r.raise_for_status()
    except Exception as e:
        log.warning(f'[kde_product] XLSX download failed: {e}')
        return []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    except Exception as e:
        log.warning(f'[kde_product] XLSX parse error: {e}')
        return []

    perf_points = []

    for sheet in wb.worksheets:
        log.info(f'[kde_product] Sheet: {sheet.title} ({sheet.max_row} rows x {sheet.max_column} cols)')
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 3:
            continue

        # KDE XLSX has a 2-row header:
        # Row 1: MOTOR VERSION | VOLTAGE | PROPELLER | THROTTLE | AMPERAGE | POWER INPUT | | THRUST OUTPUT | | | EST THRUST | | | RPM | TORQUE | Efficiency
        # Row 2: (units)      | LiPo[V] | SIZE      | RANGE    | [A]      | [W]  | [HP] | [g] | [N] | [lb] | [g] | [N] | [lb] | [rev/min] | [Nm] | [g/W]
        # Row 3+: data
        #
        # Fixed column indices (0-based):
        COL = {
            'motor_ver':   0,
            'voltage':     1,
            'prop':        2,
            'throttle':    3,
            'current':     4,
            'power_w':     5,
            'power_hp':    6,
            'thrust_g':    7,
            'thrust_n':    8,
            'thrust_lb':   9,
            'sea_thrust_g':10,
            'sea_thrust_n':11,
            'sea_thrust_lb':12,
            'rpm':         13,
            'torque':      14,
            'efficiency':  15,
        }

        def safe_float(v):
            try: return float(v) if v is not None else None
            except: return None

        def get_col(row, key):
            idx = COL.get(key, -1)
            return row[idx] if 0 <= idx < len(row) else None

        # Verify it looks like a KDE performance sheet
        row0 = rows[0]
        if not any(str(c or '').upper() in ('MOTOR VERSION', 'THROTTLE', 'AMPERAGE') for c in row0):
            log.info(f'[kde_product] Sheet {sheet.title} does not look like performance data, skipping')
            continue

        # Parse from row 3 onwards (index 2)
        motor_ver, voltage, prop = '', '', ''

        for row in rows[2:]:
            if all(c is None for c in row):
                continue

            # Carry forward rowspan values
            if get_col(row, 'motor_ver') is not None:
                motor_ver = str(get_col(row, 'motor_ver')).strip()
            if get_col(row, 'voltage') is not None:
                voltage = str(get_col(row, 'voltage')).strip()
            if get_col(row, 'prop') is not None:
                prop = str(get_col(row, 'prop')).strip()

            throttle = get_col(row, 'throttle')
            thrust_g = get_col(row, 'thrust_g')

            if throttle is None or thrust_g is None:
                continue
            try:
                throttle = float(throttle)
                thrust_g = float(thrust_g)
            except (TypeError, ValueError):
                continue

            # Convert throttle 0.25→25%
            if 0 < throttle <= 1.0:
                throttle = round(throttle * 100, 1)

            perf_points.append({
                'label':        motor_name,
                'motor_ver':    motor_ver,
                'prop':         prop,
                'throttle':     throttle,
                'voltage':      voltage,
                'current':      safe_float(get_col(row, 'current')),
                'power':        safe_float(get_col(row, 'power_w')),
                'thrust_g':     thrust_g,
                'thrust_n':     safe_float(get_col(row, 'thrust_n')),
                'thrust_lb':    safe_float(get_col(row, 'thrust_lb')),
                'rpm':          safe_float(get_col(row, 'rpm')),
                'torque':       safe_float(get_col(row, 'torque')),
                'efficiency':   safe_float(get_col(row, 'efficiency')),
                'source':       'kdedirect',
                'source_url':   xlsx_url,
            })

    log.info(f'[kde_product] Parsed {len(perf_points)} performance points from XLSX')
    return perf_points
