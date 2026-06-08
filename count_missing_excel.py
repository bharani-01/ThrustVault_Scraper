import openpyxl

file_path = r"d:\motor data\Motor List.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    total_rows = 0
    missing_rows = 0
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # Iterate rows starting from row 3 (first product row)
        rows = list(sheet.iter_rows(min_row=3, values_only=True))
        sheet_total = 0
        sheet_missing = 0
        for r in rows:
            # Skip empty rows
            if not any(r):
                continue
            sheet_total += 1
            # Check if any key field is missing
            # Columns: 0: MOTOR, 1: Company, 2: Max Thrust, 3: Recommended ESC, 4: Recommended Propeller, 5: LINK - Motor, 6: LINK - ESC, 7: LINK - Propeller
            motor = r[0]
            if not motor or len(str(motor).strip()) < 3:
                continue
                
            is_missing = False
            # If any of: Company, Max Thrust, Rec. ESC, Rec. Propeller, Link Motor is missing/empty
            # We check if they are None or empty string
            for idx in [1, 2, 3, 4, 5]:
                val = r[idx] if idx < len(r) else None
                if val is None or str(val).strip() == "":
                    is_missing = True
                    break
            if is_missing:
                sheet_missing += 1
                
        print(f"Sheet '{sheet_name}': {sheet_total} total motors, {sheet_missing} missing some specs.")
        total_rows += sheet_total
        missing_rows += sheet_missing
        
    print(f"\nOverall: {total_rows} total motors, {missing_rows} with missing specifications.")
except Exception as e:
    print("Error:", e)
