import openpyxl

file_path = r"d:\motor data\Motor List.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    print("Sheets in workbook:", wb.sheetnames)
    
    # Let's inspect the active sheet or the first sheet
    sheet = wb.active
    print(f"\nActive Sheet Name: {sheet.title}")
    print(f"Max row: {sheet.max_row}, Max col: {sheet.max_column}")
    
    # Read first 10 rows
    print("\nFirst 10 rows:")
    for row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), 1):
        # Filter out None values or format them nicely
        row_vals = [str(val) if val is not None else "" for val in row]
        # Skip completely empty rows
        if not any(row_vals):
            continue
        print(f"Row {row_idx}: {row_vals}")
        
except Exception as e:
    print(f"Error reading Excel file: {e}")
